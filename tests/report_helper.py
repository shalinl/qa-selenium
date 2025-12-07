from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

file_name = "../reports/Book1.xlsx"

def write_result(test_id, title,module,steps,indata,eresult,aresult,status):
    date = datetime.now().strftime("%d-%m-%Y")
    time = datetime.now().strftime("%H:%M:%S")

    # If file doesn't exist, create it
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(["Test Case ID", "Title","Module", "Steps", "Test Data", "Expected Result", "Actual Result","Status"])
        wb.save(file_name)

    # Open existing file
    wb = load_workbook(file_name)
    ws = wb.active

    # Append new result
    ws.append([test_id, title,module,steps,indata,eresult,aresult,status])

    wb.save(file_name)